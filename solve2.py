import numpy as np
import pandas as pd
from scipy.interpolate import interp1d
from scipy.optimize import minimize
from scipy.stats import ttest_1samp
from pulp import *

# =========================================================================
# 全局配置
# =========================================================================
DATA_DIR = '/home/lsp/Mathematical_schrace/data/'
PHOTO_ANGLE_MIN_DIFF = 60.0  # 拍照角度差异阈值（度）

# =========================================================================
# 工具函数
# =========================================================================
def calculate_angle_diff(a1, a2):
    """计算两角度在 [-180, 180] 范围内的最小夹角差"""
    diff = np.abs(a1 - a2) % 360
    return np.where(diff > 180, 360 - diff, diff)

def estimate_initial_dt(df1, df2):
    """智能估计初始时间偏差：用两个传感器时间范围中点的差值"""
    t1_mid = (df1.iloc[:, 0].min() + df1.iloc[:, 0].max()) / 2
    t2_mid = (df2.iloc[:, 0].min() + df2.iloc[:, 0].max()) / 2
    return t1_mid - t2_mid

# =========================================================================
# 问题1：纯时间对齐（三次样条插值 + 简单平均融合）
# =========================================================================
def solve_problem_1(df1, df2):
    t1, x1, y1 = df1.iloc[:, 0].values, df1.iloc[:, 1].values, df1.iloc[:, 2].values
    t2, x2, y2 = df2.iloc[:, 0].values, df2.iloc[:, 1].values, df2.iloc[:, 2].values

    f_x2 = interp1d(t2, x2, kind='cubic', fill_value='extrapolate')
    f_y2 = interp1d(t2, y2, kind='cubic', fill_value='extrapolate')

    dt_init = estimate_initial_dt(df1, df2)

    def objective(dt):
        t2_proj = t1 - dt
        mask = (t2_proj >= t2.min()) & (t2_proj <= t2.max())
        if not np.any(mask):
            return 1e9
        return np.mean((x1[mask] - f_x2(t2_proj[mask]))**2 +
                       (y1[mask] - f_y2(t2_proj[mask]))**2)

    res = minimize(objective, [dt_init], method='Nelder-Mead',
                   options={'maxiter': 500, 'xatol': 1e-8, 'fatol': 1e-8})
    dt_opt = res.x[0]

    t_start = max(t1.min(), t2.min() + dt_opt)
    t_end = min(t1.max(), t2.max() + dt_opt)
    t_10hz = np.arange(t_start, t_end, 0.1)

    f_x1 = interp1d(t1, x1, kind='cubic', fill_value='extrapolate')
    f_y1 = interp1d(t1, y1, kind='cubic', fill_value='extrapolate')

    x_fuse = (f_x1(t_10hz) + f_x2(t_10hz - dt_opt)) / 2
    y_fuse = (f_y1(t_10hz) + f_y2(t_10hz - dt_opt)) / 2

    traj = pd.DataFrame({'时间(s)': t_10hz, 'X坐标(m)': x_fuse, 'Y坐标(m)': y_fuse})
    return dt_opt, traj

# =========================================================================
# 时空联合对齐
# =========================================================================
def align_spatiotemporal(df1, df2, estimate_bias=True):
    t1, x1, y1 = df1.iloc[:, 0].values, df1.iloc[:, 1].values, df1.iloc[:, 2].values
    t2, x2, y2 = df2.iloc[:, 0].values, df2.iloc[:, 1].values, df2.iloc[:, 2].values

    f_x2 = interp1d(t2, x2, kind='linear', fill_value='extrapolate')
    f_y2 = interp1d(t2, y2, kind='linear', fill_value='extrapolate')

    dt_init = estimate_initial_dt(df1, df2)

    def objective(params):
        dt = params[0]
        dx = params[1] if estimate_bias else 0.0
        dy = params[2] if estimate_bias else 0.0
        t2_proj = t1 - dt
        mask = (t2_proj >= t2.min()) & (t2_proj <= t2.max())
        if not np.any(mask):
            return 1e9
        pred_x = f_x2(t2_proj[mask]) + dx
        pred_y = f_y2(t2_proj[mask]) + dy
        return np.mean((x1[mask] - pred_x)**2 + (y1[mask] - pred_y)**2)

    if estimate_bias:
        init = [dt_init, np.mean(x1) - np.mean(x2), np.mean(y1) - np.mean(y2)]
    else:
        init = [dt_init]

    res = minimize(objective, init, method='Nelder-Mead',
                   options={'maxiter': 1000, 'xatol': 1e-8, 'fatol': 1e-8})

    if estimate_bias:
        return res.x[0], res.x[1], res.x[2]
    return res.x[0], 0.0, 0.0

# =========================================================================
# 自适应观测噪声估计
# =========================================================================
def estimate_measurement_noise(df1, df2, dt, dx, dy):
    t1 = df1.iloc[:, 0].values
    t2 = df2.iloc[:, 0].values + dt

    fx = interp1d(t2, df2.iloc[:, 1].values + dx, fill_value='extrapolate', kind='linear')
    fy = interp1d(t2, df2.iloc[:, 2].values + dy, fill_value='extrapolate', kind='linear')

    mask = (t1 >= t2.min()) & (t1 <= t2.max())
    ex = df1.iloc[:, 1].values[mask] - fx(t1[mask])
    ey = df1.iloc[:, 2].values[mask] - fy(t1[mask])

    return np.diag([max(np.var(ex), 1e-4), max(np.var(ey), 1e-4)])

# =========================================================================
# 异步多速率扩展卡尔曼滤波（6维状态：[x, y, vx, vy, ax, ay]）
# =========================================================================
def asynchronous_kalman_filter(df1, df2, dt_opt, dx_opt, dy_opt, R1, R2):
    df2_aligned = df2.copy()
    df2_aligned.iloc[:, 0] += dt_opt
    df2_aligned.iloc[:, 1] += dx_opt
    df2_aligned.iloc[:, 2] += dy_opt

    t_start = max(df1.iloc[:, 0].min(), df2_aligned.iloc[:, 0].min())
    t_end = min(df1.iloc[:, 0].max(), df2_aligned.iloc[:, 0].max())
    t_grid = np.arange(t_start, t_end + 1e-5, 0.1)

    events = []
    for t in t_grid:
        events.append({'t': t, 'type': 'grid'})
    for _, row in df1.iterrows():
        events.append({'t': row.iloc[0], 'type': 'w1', 'z': row.iloc[1:3].values})
    for _, row in df2_aligned.iterrows():
        events.append({'t': row.iloc[0], 'type': 'w2', 'z': row.iloc[1:3].values})
    events.sort(key=lambda e: e['t'])

    X = np.array([df1.iloc[0, 1], df1.iloc[0, 2], 0.0, 0.0, 0.0, 0.0])
    P = np.eye(6) * 1.0
    Q_base = np.eye(6) * 0.05
    H = np.zeros((2, 6))
    H[0, 0] = 1
    H[1, 1] = 1

    current_t = events[0]['t']
    output = []

    for event in events:
        t_next = event['t']
        delta_t = t_next - current_t

        if delta_t > 0:
            F = np.eye(6)
            F[0, 2], F[1, 3] = delta_t, delta_t
            F[0, 4], F[1, 5] = 0.5 * delta_t**2, 0.5 * delta_t**2
            F[2, 4], F[3, 5] = delta_t, delta_t
            X = F @ X
            P = F @ P @ F.T + Q_base * delta_t
            current_t = t_next

        if event['type'] in ['w1', 'w2']:
            R = R1 if event['type'] == 'w1' else R2
            Y = event['z'] - H @ X
            S = H @ P @ H.T + R
            K = P @ H.T @ np.linalg.inv(S)
            X = X + K @ Y
            P = (np.eye(6) - K @ H) @ P
        elif event['type'] == 'grid':
            output.append({
                '时间(s)': t_next, 'X坐标(m)': X[0], 'Y坐标(m)': X[1],
                'Vx(m/s)': X[2], 'Vy(m/s)': X[3],
                'Ax(m/s^2)': X[4], 'Ay(m/s^2)': X[5]
            })

    res_df = pd.DataFrame(output).drop_duplicates(subset=['时间(s)'], keep='last')
    return res_df

# =========================================================================
# 问题2：含随机噪声+固定系统偏差
# =========================================================================
def solve_problem_2(df1, df2):
    dt, dx, dy = align_spatiotemporal(df1, df2, estimate_bias=True)
    R1 = estimate_measurement_noise(df1, df2, dt, dx, dy)
    R2 = R1.copy()
    traj = asynchronous_kalman_filter(df1, df2, dt, dx, dy, R1, R2)
    return dt, dx, dy, traj

# =========================================================================
# 问题3：先纯时间对齐 → t检验 → 再决定是否引入空间偏差
# =========================================================================
def solve_problem_3(df1, df2):
    # Step 1: 仅做时间对齐
    dt_only, _, _ = align_spatiotemporal(df1, df2, estimate_bias=False)
    print(f"  [Step1] 纯时间对齐结果: dt = {dt_only:.4f}s")

    # Step 2: 计算空间残差
    t1 = df1.iloc[:, 0].values
    x1, y1 = df1.iloc[:, 1].values, df1.iloc[:, 2].values
    t2_raw = df2.iloc[:, 0].values
    x2_raw, y2_raw = df2.iloc[:, 1].values, df2.iloc[:, 2].values

    t2_aligned = t2_raw + dt_only
    f_x2 = interp1d(t2_aligned, x2_raw, kind='linear', fill_value='extrapolate')
    f_y2 = interp1d(t2_aligned, y2_raw, kind='linear', fill_value='extrapolate')

    mask = (t1 >= t2_aligned.min()) & (t1 <= t2_aligned.max())
    n_overlap = np.sum(mask)
    print(f"  [Step2] 重叠样本数: {n_overlap}")

    if n_overlap < 10:
        print("  ⚠ 样本过少，强制设为有系统偏差模式")
        dt_opt, dx_opt, dy_opt = align_spatiotemporal(df1, df2, estimate_bias=True)
    else:
        dx_residual = x1[mask] - f_x2(t1[mask])
        dy_residual = y1[mask] - f_y2(t1[mask])

        # Step 3: 双尾t检验
        px = ttest_1samp(dx_residual, 0).pvalue
        py = ttest_1samp(dy_residual, 0).pvalue
        print(f"  [Step3] X方向 p={px:.5f}, 均值={np.mean(dx_residual):.4f}m, std={np.std(dx_residual):.4f}m")
        print(f"          Y方向 p={py:.5f}, 均值={np.mean(dy_residual):.4f}m, std={np.std(dy_residual):.4f}m")

        alpha = 0.05
        if px < alpha or py < alpha:
            print(f"  [Step4] 存在显著系统偏差 (p<{alpha})，联合对齐...")
            dt_opt, dx_opt, dy_opt = align_spatiotemporal(df1, df2, estimate_bias=True)
        else:
            print(f"  [Step4] 系统偏差不显著，dx=0, dy=0")
            dt_opt, dx_opt, dy_opt = dt_only, 0.0, 0.0

    # Step 5: 卡尔曼滤波
    R1 = estimate_measurement_noise(df1, df2, dt_opt, dx_opt, dy_opt)
    R2 = R1.copy()
    traj = asynchronous_kalman_filter(df1, df2, dt_opt, dx_opt, dy_opt, R1, R2)
    return dt_opt, dx_opt, dy_opt, traj

# =========================================================================
# 问题4：任务排程 ILP 优化
# =========================================================================
def solve_problem_4(trajectory_df, df_shoot, df_photo):
    t = trajectory_df['时间(s)'].values
    x = trajectory_df['X坐标(m)'].values
    y = trajectory_df['Y坐标(m)'].values
    v = np.sqrt(trajectory_df['Vx(m/s)'].values**2 + trajectory_df['Vy(m/s)'].values**2)
    a = np.sqrt(trajectory_df['Ax(m/s^2)'].values**2 + trajectory_df['Ay(m/s^2)'].values**2)

    SHOOT_STEPS = 15   # 1.5s @ 10Hz
    PHOTO_STEPS = 5    # 0.5s @ 10Hz

    candidate_windows = []

    # 射击候选窗口
    for _, target in df_shoot.iterrows():
        tid, tx, ty = target['编号'], target['X坐标(m)'], target['Y坐标(m)']
        dists = np.sqrt((x - tx)**2 + (y - ty)**2)
        valid = (dists >= 5) & (dists <= 30) & (v <= 2.0) & (a <= 1.5)
        for i in range(len(t) - SHOOT_STEPS):
            if np.all(valid[i:i + SHOOT_STEPS + 1]):
                exec_t = t[i + SHOOT_STEPS]
                angle = np.arctan2(ty - y[i + SHOOT_STEPS], tx - x[i + SHOOT_STEPS]) * 180 / np.pi
                candidate_windows.append({
                    'target_id': tid, 'type': '射击',
                    'start_t': t[i], 'end_t': exec_t, 'angle': angle
                })

    # 拍照候选窗口
    for _, target in df_photo.iterrows():
        tid, tx, ty = target['编号'], target['X坐标(m)'], target['Y坐标(m)']
        dists = np.sqrt((x - tx)**2 + (y - ty)**2)
        valid = (dists >= 10) & (dists <= 40) & (v <= 1.5) & (a <= 1.5)
        for i in range(len(t) - PHOTO_STEPS):
            if np.all(valid[i:i + PHOTO_STEPS + 1]):
                exec_t = t[i + PHOTO_STEPS]
                angle = np.arctan2(ty - y[i + PHOTO_STEPS], tx - x[i + PHOTO_STEPS]) * 180 / np.pi
                candidate_windows.append({
                    'target_id': tid, 'type': '拍照',
                    'start_t': t[i], 'end_t': exec_t, 'angle': angle
                })

    if not candidate_windows:
        print("⚠ 未找到任何可行任务窗口！")
        return pd.DataFrame()

    print(f"  共生成 {len(candidate_windows)} 个候选窗口")

    # ILP
    prob = LpProblem("Robot_Task_Optimization", LpMaximize)
    N = len(candidate_windows)
    x_var = {i: LpVariable(f"x_{i}", cat='Binary') for i in range(N)}

    prob += lpSum([(0.85 if candidate_windows[i]['type'] == '射击' else 1.0) * x_var[i]
                   for i in range(N)])

    # 时间冲突约束
    all_times = sorted(set([w['start_t'] for w in candidate_windows] +
                           [w['end_t'] for w in candidate_windows]))
    for k in range(len(all_times) - 1):
        t_mid = (all_times[k] + all_times[k + 1]) / 2
        active = [i for i, w in enumerate(candidate_windows)
                  if w['start_t'] <= t_mid <= w['end_t']]
        if len(active) > 1:
            prob += lpSum([x_var[i] for i in active]) <= 1

    # 拍照角度差异约束
    photo_by_target = {}
    for i, w in enumerate(candidate_windows):
        if w['type'] == '拍照':
            photo_by_target.setdefault(w['target_id'], []).append(i)

    angle_cnstr = 0
    for tid, indices in photo_by_target.items():
        for a in range(len(indices)):
            for b in range(a + 1, len(indices)):
                i, j = indices[a], indices[b]
                diff = calculate_angle_diff(candidate_windows[i]['angle'],
                                            candidate_windows[j]['angle'])
                if diff < PHOTO_ANGLE_MIN_DIFF:
                    prob += x_var[i] + x_var[j] <= 1
                    angle_cnstr += 1
    print(f"  角度差异约束（<{PHOTO_ANGLE_MIN_DIFF}°）: {angle_cnstr} 条")

    prob.solve(PULP_CBC_CMD(msg=False))
    print(f"  求解状态: {LpStatus[prob.status]}")

    selected = []
    for i in range(N):
        if pulp.value(x_var[i]) > 0.5:
            selected.append(candidate_windows[i])
    selected.sort(key=lambda w: w['start_t'])
    return pd.DataFrame(selected)

# =========================================================================
# 输出 result.xlsx
# =========================================================================
def write_result_xlsx(schedule_df, template_path, output_path):
    from openpyxl import load_workbook
    wb = load_workbook(template_path)
    ws = wb.active
    start_row = 4
    for idx, (_, row_data) in enumerate(schedule_df.iterrows()):
        r = start_row + idx
        ws.cell(row=r, column=1, value=idx + 1)
        ws.cell(row=r, column=2, value=row_data['target_id'])
        ws.cell(row=r, column=3, value=row_data['type'])
        ws.cell(row=r, column=4, value=round(row_data['start_t'], 2))
        ws.cell(row=r, column=5, value=round(row_data['end_t'], 2))
    wb.save(output_path)
    print(f"✅ 结果已写入: {output_path}")

# =========================================================================
# 主程序
# =========================================================================
if __name__ == '__main__':
    DATA = DATA_DIR
    print("=" * 60)
    print("  多源融合机器人定位及任务优化")
    print("=" * 60)

    print("\n📂 加载数据...")
    f1_p1 = pd.read_excel(DATA + '附件1.xlsx', sheet_name='方式1(4Hz)')
    f2_p1 = pd.read_excel(DATA + '附件1.xlsx', sheet_name='方式2(5Hz)')
    f1_p2 = pd.read_excel(DATA + '附件2.xlsx', sheet_name='方式1(4Hz)')
    f2_p2 = pd.read_excel(DATA + '附件2.xlsx', sheet_name='方式2(5Hz)')
    f1_p3 = pd.read_excel(DATA + '附件3.xlsx', sheet_name='方式1(4Hz)')
    f2_p3 = pd.read_excel(DATA + '附件3.xlsx', sheet_name='方式2(5Hz)')
    df_shoot = pd.read_excel(DATA + '附件4.xlsx', sheet_name='射击目标')
    df_photo = pd.read_excel(DATA + '附件4.xlsx', sheet_name='拍照目标')

    # 问题1
    print("\n" + "=" * 40)
    print("[问题1] 无噪声，仅时间对齐")
    dt1, traj1 = solve_problem_1(f1_p1, f2_p1)
    print(f"  时间偏差 dt = {dt1:.4f}s, 10Hz轨迹: {len(traj1)}点")
    traj1.to_csv(DATA + '问题1_10Hz轨迹结果.csv', index=False)

    # 问题2
    print("\n" + "=" * 40)
    print("[问题2] 含随机噪声+固定系统偏差")
    dt2, dx2, dy2, traj2 = solve_problem_2(f1_p2, f2_p2)
    print(f"  dt = {dt2:.4f}s, dx = {dx2:.4f}m, dy = {dy2:.4f}m, 轨迹: {len(traj2)}点")
    traj2.to_csv(DATA + '问题2_10Hz轨迹结果.csv', index=False)

    # 问题3
    print("\n" + "=" * 40)
    print("[问题3] 实际数据，判断系统偏差 → 对齐融合")
    dt3, dx3, dy3, traj3 = solve_problem_3(f1_p3, f2_p3)
    print(f"  最终: dt = {dt3:.4f}s, dx = {dx3:.4f}m, dy = {dy3:.4f}m, 轨迹: {len(traj3)}点")
    traj3.to_csv(DATA + '问题3_10Hz轨迹结果.csv', index=False)

    # 问题4
    print("\n" + "=" * 40)
    print("[问题4] 任务排程优化")
    schedule_df = solve_problem_4(traj3, df_shoot, df_photo)

    if not schedule_df.empty:
        s_cnt = (schedule_df['type'] == '射击').sum()
        p_cnt = (schedule_df['type'] == '拍照').sum()
        print(f"  射击: {s_cnt}, 拍照: {p_cnt}, 总计: {len(schedule_df)}")
        print(schedule_df[['target_id', 'type', 'start_t', 'end_t', 'angle']].head(20).to_string())
        write_result_xlsx(schedule_df, DATA + 'result.xlsx', DATA + 'result.xlsx')
    else:
        print("  ⚠ 无可行排程结果！")

    print("\n" + "=" * 60)
    print("  全部求解完成！")
    print("=" * 60)
